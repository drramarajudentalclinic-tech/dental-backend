"""
routes/other_expenses.py
────────────────────────
Blueprint for Other Expenses.

Every expense is stored in:
  - SQLite (OtherExpense table) — for fast querying / editing / deleting
  - Excel  (receipts/other expenses/<filename>.xlsx) — one file per calendar
    month, auto-created on first write, auto-named:
        "Other Exp March/2026.xlsx"

Excel columns  (A–F):
  ID | Date | Paid To (Type) | Name | Amount (₹) | Description

On EDIT   → row is found by ID column and updated in-place.
On DELETE → row is found by ID column and removed; rows below shift up.
"""

import os
from datetime import datetime, date

from flask        import Blueprint, request, jsonify
from flask_sqlalchemy import SQLAlchemy

# ── shared db instance (same as rest of app) ─────────────────────────────────
from database import db

# ── openpyxl ─────────────────────────────────────────────────────────────────
from openpyxl                        import Workbook, load_workbook
from openpyxl.styles                 import (Font, PatternFill, Alignment,
                                              Border, Side)
from openpyxl.utils                  import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Blueprint
# ─────────────────────────────────────────────────────────────────────────────
other_expenses_bp = Blueprint("other_expenses", __name__, url_prefix="/api")

# ─────────────────────────────────────────────────────────────────────────────
# SQLAlchemy Model
# ─────────────────────────────────────────────────────────────────────────────
class OtherExpense(db.Model):
    __tablename__ = "other_expenses"

    id          = db.Column(db.Integer,  primary_key=True)
    date        = db.Column(db.String(20),  nullable=False)   # "YYYY-MM-DD"
    type        = db.Column(db.String(20),  nullable=False)   # "Dr"|"Company"|"Other"
    party_name  = db.Column(db.String(200), nullable=False)
    amount      = db.Column(db.Float,       nullable=False)
    description = db.Column(db.Text,        nullable=True, default="")
    excel_file  = db.Column(db.String(200), nullable=True)    # e.g. "Other Exp March/2026"
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            "id":          self.id,
            "date":        self.date,
            "type":        self.type,
            "party_name":  self.party_name,
            "amount":      self.amount,
            "description": self.description or "",
            "excel_file":  self.excel_file or "",
            "created_at":  self.created_at.isoformat() if self.created_at else None,
        }


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

# Base folder: receipts/other expenses/
RECEIPTS_BASE = os.path.join("receipts", "other expenses")


def _excel_path(expense_date: str) -> str:
    """Return the full path to the Excel file for this expense's month.

    expense_date: "YYYY-MM-DD"
    Returns:  receipts/other expenses/Other Exp March_2026.xlsx
    """
    d = datetime.strptime(expense_date, "%Y-%m-%d")
    month_name = d.strftime("%B")   # "March"
    year       = d.strftime("%Y")   # "2026"
    filename   = f"Other Exp {month_name}_{year}.xlsx"
    return os.path.join(RECEIPTS_BASE, filename)


# Styled header row
_HEADER_FILL  = PatternFill("solid", start_color="1B3A5C", end_color="1B3A5C")
_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_TOTAL_FILL   = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")
_TOTAL_FONT   = Font(name="Arial", bold=True, color="78350F", size=11)
_THIN_BORDER  = Border(
    left   = Side(style="thin", color="D1D5DB"),
    right  = Side(style="thin", color="D1D5DB"),
    top    = Side(style="thin", color="D1D5DB"),
    bottom = Side(style="thin", color="D1D5DB"),
)
_DATA_FONT    = Font(name="Arial", size=10)
_COLS         = ["ID", "Date", "Paid To (Type)", "Name", "Amount (₹)", "Description"]
_COL_WIDTHS   = [8, 14, 16, 28, 16, 42]


def _ensure_excel(path: str):
    """Create Excel file with header row if it doesn't exist yet."""
    if os.path.exists(path):
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()
    ws = wb.active

    # Title row (row 1)
    fname  = os.path.splitext(os.path.basename(path))[0]   # "Other Exp March_2026"
    title  = fname.replace("_", "/")
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = f"💸  {title}  —  Other Expenses"
    title_cell.font      = Font(name="Arial", bold=True, color="78350F", size=13)
    title_cell.fill      = PatternFill("solid", start_color="FFF7ED", end_color="FFF7ED")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row (row 2)
    for col_idx, (header, width) in enumerate(zip(_COLS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    wb.save(path)


def _append_row(path: str, exp: OtherExpense):
    """Append a single data row. Rebuilds TOTAL row at bottom."""
    _ensure_excel(path)
    wb = load_workbook(path)
    ws = wb.active

    # Find first empty data row (after header rows 1 & 2, skip total row if any)
    last_data_row = 2
    for row in ws.iter_rows(min_row=3):
        first_val = row[0].value
        if first_val is None or str(first_val).strip() == "":
            break
        if str(first_val).strip().upper() == "TOTAL":
            break
        last_data_row = row[0].row

    new_row = last_data_row + 1

    # Remove old TOTAL row if it exists at new_row position
    if ws.cell(row=new_row, column=1).value is not None and \
       str(ws.cell(row=new_row, column=1).value).strip().upper() == "TOTAL":
        ws.delete_rows(new_row)

    # Write data
    row_data = [
        exp.id,
        exp.date,
        exp.type,
        exp.party_name,
        exp.amount,
        exp.description or "",
    ]
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=new_row, column=col_idx, value=val)
        cell.font      = _DATA_FONT
        cell.border    = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if col_idx == 5:   # amount column — right-align, number format
            cell.alignment   = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0.00'

    ws.row_dimensions[new_row].height = 18

    # Rebuild TOTAL row at the end
    _write_total_row(ws)

    wb.save(path)


def _write_total_row(ws):
    """Add / update the TOTAL row at the bottom of data."""
    # Find last data row (excluding old total)
    last_data_row = 2
    for row in ws.iter_rows(min_row=3):
        cell_val = row[0].value
        if cell_val is None or str(cell_val).strip() == "":
            break
        if str(cell_val).strip().upper() == "TOTAL":
            continue
        last_data_row = row[0].row

    total_row = last_data_row + 1

    # Clear any existing total row
    for col_idx in range(1, 7):
        ws.cell(row=total_row, column=col_idx).value = None

    # Write TOTAL
    ws.cell(row=total_row, column=1, value="TOTAL").font  = _TOTAL_FONT
    ws.cell(row=total_row, column=1).fill                 = _TOTAL_FILL
    ws.cell(row=total_row, column=1).alignment            = Alignment(horizontal="right", vertical="center")

    # SUM formula over amount column (E), rows 3 to last_data_row
    if last_data_row >= 3:
        formula = f"=SUM(E3:E{last_data_row})"
    else:
        formula = 0

    total_amount_cell = ws.cell(row=total_row, column=5, value=formula)
    total_amount_cell.font         = _TOTAL_FONT
    total_amount_cell.fill         = _TOTAL_FILL
    total_amount_cell.alignment    = Alignment(horizontal="right", vertical="center")
    total_amount_cell.number_format = '#,##0.00'

    for col_idx in [2, 3, 4, 6]:
        c = ws.cell(row=total_row, column=col_idx)
        c.fill      = _TOTAL_FILL
        c.border    = _THIN_BORDER

    for col_idx in [1, 5]:
        ws.cell(row=total_row, column=col_idx).border = _THIN_BORDER

    ws.row_dimensions[total_row].height = 20


def _update_row_in_excel(path: str, exp: OtherExpense):
    """Find row by ID (column A) and update it in place."""
    if not os.path.exists(path):
        # File doesn't exist yet — just append
        _append_row(path, exp)
        return

    wb = load_workbook(path)
    ws = wb.active
    updated = False

    for row in ws.iter_rows(min_row=3):
        if row[0].value == exp.id:
            row_data = [exp.id, exp.date, exp.type,
                        exp.party_name, exp.amount, exp.description or ""]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row[0].row, column=col_idx, value=val)
                cell.font      = _DATA_FONT
                cell.border    = _THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if col_idx == 5:
                    cell.alignment    = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00'
            updated = True
            break

    if not updated:
        # Row not found (maybe month changed) — append to new file
        wb.close()
        _append_row(path, exp)
        return

    _write_total_row(ws)
    wb.save(path)


def _delete_row_from_excel(path: str, expense_id: int):
    """Remove the row matching expense_id (column A) from the Excel file."""
    if not os.path.exists(path):
        return

    wb = load_workbook(path)
    ws = wb.active

    target_row = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value == expense_id:
            target_row = row[0].row
            break

    if target_row:
        ws.delete_rows(target_row)
        # Rebuild TOTAL row
        _write_total_row(ws)

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Migration helper (called from app.py)
# ─────────────────────────────────────────────────────────────────────────────
def run_other_expense_migrations(app):
    """Create table if not yet present. Safe to call multiple times."""
    with app.app_context():
        db.create_all()


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@other_expenses_bp.route("/other-expenses", methods=["GET"])
def list_expenses():
    """Return all expenses, newest first."""
    rows = OtherExpense.query.order_by(OtherExpense.date.desc(),
                                       OtherExpense.id.desc()).all()
    return jsonify([r.to_dict() for r in rows]), 200


@other_expenses_bp.route("/other-expenses", methods=["POST"])
def create_expense():
    """Create expense in DB and append to Excel."""
    data = request.get_json(force=True)

    required = ("type", "party_name", "amount", "date")
    missing  = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    try:
        amount = float(data["amount"])
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid amount"}), 400

    exp = OtherExpense(
        date        = data["date"],
        type        = data["type"],
        party_name  = data["party_name"],
        amount      = amount,
        description = data.get("description", ""),
        excel_file  = data.get("excel_file", ""),
    )
    db.session.add(exp)
    db.session.commit()   # commit first so exp.id is populated

    # Write to Excel
    try:
        path = _excel_path(exp.date)
        _append_row(path, exp)
    except Exception as e:
        import traceback
        print(f"[OtherExpenses] Excel append failed: {e}")
        traceback.print_exc()

    return jsonify(exp.to_dict()), 201


@other_expenses_bp.route("/other-expenses/<int:expense_id>", methods=["PUT"])
def update_expense(expense_id):
    """Update expense in DB and update the Excel row."""
    exp = OtherExpense.query.get_or_404(expense_id)
    data = request.get_json(force=True)

    old_date = exp.date   # need the old date to know which Excel file had this row

    # Update fields
    if "date"        in data: exp.date        = data["date"]
    if "type"        in data: exp.type        = data["type"]
    if "party_name"  in data: exp.party_name  = data["party_name"]
    if "description" in data: exp.description = data["description"]
    if "excel_file"  in data: exp.excel_file  = data["excel_file"]
    if "amount"      in data:
        try:
            exp.amount = float(data["amount"])
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid amount"}), 400

    db.session.commit()

    # Update Excel — handle month change (delete from old file, add to new)
    try:
        old_path = _excel_path(old_date)
        new_path = _excel_path(exp.date)

        if old_path != new_path:
            # Month changed: remove from old sheet, append to new
            _delete_row_from_excel(old_path, exp.id)
            _append_row(new_path, exp)
        else:
            _update_row_in_excel(new_path, exp)
    except Exception as e:
        import traceback
        print(f"[OtherExpenses] Excel update failed: {e}")
        traceback.print_exc()

    return jsonify(exp.to_dict()), 200


@other_expenses_bp.route("/other-expenses/<int:expense_id>", methods=["DELETE"])
def delete_expense(expense_id):
    """Delete expense from DB and remove its row from Excel."""
    exp = OtherExpense.query.get_or_404(expense_id)

    # Remove from Excel first (while we still have the date)
    try:
        path = _excel_path(exp.date)
        _delete_row_from_excel(path, exp.id)
    except Exception as e:
        import traceback
        print(f"[OtherExpenses] Excel delete failed: {e}")
        traceback.print_exc()

    db.session.delete(exp)
    db.session.commit()

    return jsonify({"message": "Deleted", "id": expense_id}), 200