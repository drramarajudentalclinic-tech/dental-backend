import os
from openpyxl import Workbook, load_workbook
from utils.folders import get_excel_path


def init_excel(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Date",
            "Receipt Number",
            "Patient Name",
            "Case Number",
            "Mobile",
            "Visit ID",
            "Treatment",
            "Paid Amount",
            "Payment Method",
            "Status"
        ])
        wb.save(path)

def add_row(path, row):
    wb = load_workbook(path)
    ws = wb.active
    ws.append(row)
    wb.save(path)

def replace_with_deleted(path, receipt_number):
    wb = load_workbook(path)
    ws = wb.active

    for r in ws.iter_rows(min_row=2):
        if str(r[1].value) == str(receipt_number):
            r[0].value = ""
            r[2].value = "DELETED"
            for i in range(3, 9):
                r[i].value = ""
            r[9].value = "DELETED"
            break

    wb.save(path)